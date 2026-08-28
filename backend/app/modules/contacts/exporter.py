from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.modules.contacts.schemas import NewContactItem


def export_new_contacts_xlsx(
    rows: list[NewContactItem], date_from: date, date_to: date
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Новые обращения"
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells("A1:H1")
    sheet["A1"] = "Новые обращения"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1D3B31")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet["A2"] = "Период"
    sheet["B2"] = f"{date_from.isoformat()} — {date_to.isoformat()}"
    sheet["A3"] = "Количество"
    sheet["B3"] = len(rows)

    headers = [
        "Номер телефона",
        "Первое обращение",
        "Источник",
        "Последнее обращение",
        "Всего контактов",
        "Звонков",
        "Сообщений WhatsApp",
        "Статус сверки с 1С",
    ]
    header_row = 5
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F725B")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 32

    source_labels = {"kcell": "Kcell", "whatsapp": "WhatsApp"}
    for item in rows:
        first = item.first_contact_at.replace(tzinfo=None)
        last = item.last_contact_at.replace(tzinfo=None)
        sheet.append([
            item.phone_number or "Номер недоступен",
            first,
            source_labels.get(item.source, item.source),
            last,
            item.inbound_count,
            item.call_count,
            item.message_count,
            "Не найден в 1С",
        ])

    if rows:
        table = Table(displayName="NewContacts", ref=f"A{header_row}:H{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    thin = Side(style="thin", color="DDE1DA")
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
    for column in ("B", "D"):
        for cell in sheet[column][header_row:]:
            cell.number_format = "yyyy-mm-dd hh:mm"
    for column in ("E", "F", "G"):
        for cell in sheet[column][header_row:]:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right", vertical="center")

    widths = {"A": 22, "B": 21, "C": 15, "D": 21, "E": 18, "F": 12, "G": 22, "H": 22}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:H{sheet.max_row}"
    sheet.print_title_rows = f"{header_row}:{header_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
