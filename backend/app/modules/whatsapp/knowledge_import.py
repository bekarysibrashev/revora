from dataclasses import dataclass
from io import BytesIO
import re

from openpyxl import load_workbook


@dataclass(frozen=True)
class ImportedKnowledge:
    category: str
    title: str
    content_ru: str | None
    content_kk: str | None
    risk_level: str
    source: str


PROMOTIONAL_MARKERS = (
    "РАССЫЛ",
    "ВОЗВРАТ",
    "ТРИГГЕР",
    "АКЦИ",
    "СКИДК",
    "ПРОФИЛАКСИС",
    "АЛГОРИТМ ДЕЙСТВИЙ",
)
EXCLUDED_SHEET_PATTERN = re.compile(r"^ЛИСТ\s*\d+$", re.IGNORECASE)


def import_knowledge_workbook(data: bytes, filename: str) -> list[ImportedKnowledge]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    result: list[ImportedKnowledge] = []
    for sheet in workbook.worksheets:
        category = str(sheet.title).strip()
        if EXCLUDED_SHEET_PATTERN.fullmatch(category):
            continue
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = [_clean(cell) for cell in row]
            texts = [cell for cell in cells if cell]
            if len(texts) < 2 or _looks_like_header(texts):
                continue
            title_index = _title_index(texts)
            title = texts[title_index][:300]
            answers = [
                value for index, value in enumerate(texts)
                if index != title_index and len(value) >= 20
            ]
            if not answers:
                continue
            content_ru = answers[0]
            content_kk = answers[1] if len(answers) > 1 and _looks_kazakh(answers[1]) else None
            risk_text = f"{category} {title} {content_ru}".upper()
            risk = (
                "human_only"
                if any(marker in risk_text for marker in PROMOTIONAL_MARKERS)
                else "review"
            )
            result.append(
                ImportedKnowledge(
                    category=category[:120],
                    title=title,
                    content_ru=content_ru,
                    content_kk=content_kk,
                    risk_level=risk,
                    source=f"{filename}:{category}!{row_number}",
                )
            )
    return result


def _clean(value: object) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text if len(text) >= 2 else ""


def _looks_like_header(values: list[str]) -> bool:
    joined = " ".join(values).lower()
    return (
        len(values) >= 3
        and "шаблон" in joined
        and ("услуг" in joined or "сегмент" in joined)
    )


def _title_index(values: list[str]) -> int:
    for index, value in enumerate(values):
        if len(value) <= 300 and (
            "?" in value
            or value.upper() == value
            or len(value.split()) <= 12
        ):
            return index
    return 0


def _looks_kazakh(value: str) -> bool:
    return any(character in value.lower() for character in "әіңғүұқөһ")
