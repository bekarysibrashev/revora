"""Safe CSV/XLSX adapter that emits untouched, JSON-compatible source rows."""

from collections.abc import AsyncIterator, Iterable
import csv
from datetime import date, datetime
from decimal import Decimal
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from app.modules.integrations.adapter import SourceRecord


class UnsupportedTabularFile(ValueError):
    pass


class InvalidTabularFile(ValueError):
    pass


class TabularFileAdapter:
    """Read one tabular file without embedding clinic-specific mapping rules."""

    MAX_ROWS = 200_000
    MAX_COLUMNS = 500

    def __init__(
        self,
        *,
        filename: str,
        content: bytes,
        source_entity: str,
        sheet_name: str | None = None,
    ) -> None:
        self.filename = filename
        self.content = content
        self.source_entity = source_entity.strip().lower()
        self.sheet_name = sheet_name

    async def fetch(self) -> AsyncIterator[SourceRecord]:
        headers, rows = self._read()
        schema_version = sha256("\x1f".join(headers).encode("utf-8")).hexdigest()[:16]
        for row_number, values in enumerate(rows, start=2):
            if row_number > self.MAX_ROWS + 1:
                raise InvalidTabularFile(f"File exceeds {self.MAX_ROWS} data rows")
            payload = {
                header: self._json_safe(value)
                for header, value in zip(headers, values, strict=True)
            }
            if not any(value not in (None, "") for value in payload.values()):
                continue
            yield SourceRecord(
                source_entity=self.source_entity,
                payload=payload,
                external_id=str(row_number),
                schema_version=schema_version,
            )

    def _read(self) -> tuple[list[str], Iterable[tuple[object, ...]]]:
        suffix = Path(self.filename).suffix.lower()
        if suffix == ".csv":
            return self._read_csv()
        if suffix == ".xlsx":
            return self._read_xlsx()
        raise UnsupportedTabularFile("Only .csv and .xlsx files are supported")

    def _read_csv(self) -> tuple[list[str], list[tuple[object, ...]]]:
        try:
            text = self.content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = self.content.decode("cp1251")
            except UnicodeDecodeError as exc:
                raise InvalidTabularFile("CSV must use UTF-8 or Windows-1251 encoding") from exc

        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(StringIO(text), dialect))
        if not rows:
            raise InvalidTabularFile("File has no header row")
        headers = self._validate_headers(rows[0])
        normalized_rows = [self._fit_row(row, len(headers)) for row in rows[1:]]
        return headers, normalized_rows

    def _read_xlsx(self) -> tuple[list[str], Iterable[tuple[object, ...]]]:
        try:
            workbook = load_workbook(BytesIO(self.content), read_only=True, data_only=True)
        except Exception as exc:
            raise InvalidTabularFile("XLSX file cannot be read") from exc
        try:
            if self.sheet_name:
                if self.sheet_name not in workbook.sheetnames:
                    raise InvalidTabularFile(f"Worksheet '{self.sheet_name}' does not exist")
                sheet = workbook[self.sheet_name]
            else:
                sheet = workbook[workbook.sheetnames[0]]
            iterator = sheet.iter_rows(values_only=True)
            try:
                first_row = next(iterator)
            except StopIteration as exc:
                raise InvalidTabularFile("File has no header row") from exc
            headers = self._validate_headers(first_row)
            rows = [self._fit_row(row, len(headers)) for row in iterator]
            return headers, rows
        finally:
            workbook.close()

    def _validate_headers(self, values: Iterable[object]) -> list[str]:
        headers = [str(value).strip() if value is not None else "" for value in values]
        while headers and not headers[-1]:
            headers.pop()
        if not headers or any(not header for header in headers):
            raise InvalidTabularFile("Every populated column must have a header")
        if len(headers) > self.MAX_COLUMNS:
            raise InvalidTabularFile(f"File exceeds {self.MAX_COLUMNS} columns")
        normalized = [header.casefold() for header in headers]
        if len(normalized) != len(set(normalized)):
            raise InvalidTabularFile("Column headers must be unique")
        return headers

    @staticmethod
    def _fit_row(values: Iterable[object], width: int) -> tuple[object, ...]:
        row = tuple(values)
        if len(row) < width:
            return row + (None,) * (width - len(row))
        if any(value not in (None, "") for value in row[width:]):
            raise InvalidTabularFile("A data row contains values beyond the header columns")
        return row[:width]

    @staticmethod
    def _json_safe(value: object) -> object:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, Decimal):
            return str(value)
        if isinstance(value, (str, int, float, bool)) or value is None:
            return value
        return str(value)
