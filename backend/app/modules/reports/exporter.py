"""Deterministic Excel/PDF exports for financial reports."""

from io import BytesIO
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.modules.finance.schemas import CashFlowResponse, PnlResponse

ReportFormat = Literal["xlsx", "pdf"]


def export_pnl(report: PnlResponse, output_format: ReportFormat) -> bytes:
    rows = [
        ("Revenue (accrual)", report.revenue_accrual),
        ("Revenue (payment)", report.revenue_payment),
        ("Variable expenses", report.variable_expenses),
        ("Fixed expenses", report.fixed_expenses),
        ("Uncategorized expenses", report.uncategorized_expenses),
        ("Total expenses", report.total_expenses),
        ("Gross profit", report.gross_profit),
        ("EBITDA", report.ebitda),
        ("Net profit", report.net_profit),
    ]
    return _export("Profit and Loss", rows, report.meta, output_format)


def export_cashflow(report: CashFlowResponse, output_format: ReportFormat) -> bytes:
    rows = [
        ("Cash inflow", report.inflow),
        ("Cash outflow", report.outflow),
        ("Net cash flow", report.net_cash_flow),
        ("Closing balance", report.closing_balance),
    ]
    return _export("Cash Flow", rows, report.meta, output_format)


def _export(title, rows, meta, output_format: ReportFormat) -> bytes:
    if output_format == "xlsx":
        return _xlsx(title, rows, meta)
    if output_format == "pdf":
        return _pdf(title, rows, meta)
    raise ValueError("Unsupported report format")


def _xlsx(title, rows, meta) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:B1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="183B4E")
    sheet["A1"].alignment = Alignment(horizontal="left")
    sheet["A3"] = "Period"
    sheet["B3"] = f"{meta.date_from.isoformat()} - {meta.date_to.isoformat()}"
    sheet["A4"] = "Branch"
    sheet["B4"] = str(meta.branch_id) if meta.branch_id else "All branches"
    sheet["A5"] = "Data as of"
    sheet["B5"] = meta.data_as_of.isoformat() if meta.data_as_of else "No data"
    sheet.append([])
    sheet.append(["Metric", "Amount, KZT"])
    header_row = sheet.max_row
    for label, value in rows:
        sheet.append([label, float(value) if value is not None else None])
    sheet[f"A{header_row}:B{header_row}"][0][0].fill = PatternFill("solid", fgColor="DCE8EC")
    sheet[f"A{header_row}:B{header_row}"][0][1].fill = PatternFill("solid", fgColor="DCE8EC")
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="183B4E")
    thin = Side(style="thin", color="D9E0E3")
    for row in sheet.iter_rows(min_row=header_row, max_row=sheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(bottom=thin)
    for cell in sheet["B"][header_row:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
        cell.alignment = Alignment(horizontal="right")
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 24
    sheet.freeze_panes = f"A{header_row + 1}"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf(title, rows, meta) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=title,
        author="Revora",
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(title, styles["Title"]),
        Spacer(1, 4 * mm),
        Paragraph(
            f"Period: {meta.date_from.isoformat()} - {meta.date_to.isoformat()}",
            styles["Normal"],
        ),
        Paragraph(
            f"Branch: {meta.branch_id if meta.branch_id else 'All branches'}",
            styles["Normal"],
        ),
        Paragraph(
            f"Data as of: {meta.data_as_of.isoformat() if meta.data_as_of else 'No data'}",
            styles["Normal"],
        ),
        Spacer(1, 6 * mm),
    ]
    table_data = [["Metric", "Amount, KZT"]] + [
        [label, f"{value:,.2f}" if value is not None else "-"] for label, value in rows
    ]
    table = Table(table_data, colWidths=[105 * mm, 50 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#183B4E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E0E3")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8F9")]),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return output.getvalue()
