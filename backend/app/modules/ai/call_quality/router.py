from datetime import UTC, date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Annotated
from urllib.parse import unquote
from uuid import UUID, uuid4
from fastapi import APIRouter, Depends, Query, Request, Response, status
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import Settings, get_settings
from app.core.database import get_db_session
from app.core.errors import AppError
from app.modules.ai.call_quality.audio import ALLOWED_AUDIO_TYPES
from app.modules.ai.call_quality.defaults import ensure_default_rule_set
from app.modules.ai.call_quality.models import CallQualityAnalysis
from app.modules.ai.call_quality.pipeline import CallQualityPipeline
from app.modules.ai.call_quality.schemas import (
    CallAnalysisResponse, CallerContactListResponse, CallListResponse, CallQualityStatusResponse,
    ManualTestResponse, OperatorPerformanceResponse, RuleSetRequest, RuleSetResponse,
)
from app.modules.ai.call_quality.service import CallQualityService
from app.modules.auth.dependencies import CurrentUser
from app.modules.auth.models import UserRole
from app.modules.sales.models import Call

router = APIRouter(prefix="/call-quality", tags=["call-quality"])
Session = Annotated[AsyncSession, Depends(get_db_session)]
RuntimeSettings = Annotated[Settings, Depends(get_settings)]

@router.get("/status", response_model=CallQualityStatusResponse)
async def get_status(user: CurrentUser, session: Session) -> CallQualityStatusResponse:
    return await CallQualityService(session).status(user)

@router.get("/calls", response_model=CallListResponse)
async def list_calls(
    user: CurrentUser,
    session: Session,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=100),
    extension: str | None = None,
    direction: str | None = None,
    outcome: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    duration_min: int | None = Query(default=None, ge=0),
    duration_max: int | None = Query(default=None, ge=0),
    phone: str | None = None,
    sort_by: str = Query(default="started_at", pattern="^(started_at|duration|extension|outcome)$"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
) -> CallListResponse:
    if date_from and date_to and date_from > date_to:
        raise AppError("INVALID_DATE_RANGE", "date_from must not be after date_to", 422)
    return await CallQualityService(session).list_calls(
        user,
        page=page,
        page_size=page_size,
        extension=extension,
        direction=direction,
        outcome=outcome,
        date_from=date_from,
        date_to=date_to,
        duration_min=duration_min,
        duration_max=duration_max,
        phone=phone,
        sort_by=sort_by,
        sort_order=sort_order,
    )


@router.get("/contacts", response_model=CallerContactListResponse)
async def caller_contacts(
    user: CurrentUser,
    session: Session,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=5000),
    date_from: date | None = None,
    date_to: date | None = None,
    search: str | None = None,
    contact_type: str | None = Query(default=None, pattern="^(first|repeat)$"),
    sort_by: str = Query(default="last_call_at", pattern="^(last_call_at|first_call_at|call_count|duration|phone)$"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
) -> CallerContactListResponse:
    return await CallQualityService(session).caller_contacts(
        user, page=page, page_size=page_size, date_from=date_from, date_to=date_to,
        search=search, contact_type=contact_type, sort_by=sort_by, sort_order=sort_order,
    )


@router.get("/contacts/export")
async def export_caller_contacts(
    user: CurrentUser,
    session: Session,
    date_from: date | None = None,
    date_to: date | None = None,
) -> Response:
    result = await CallQualityService(session).caller_contacts(
        user, page=1, page_size=5000, date_from=date_from, date_to=date_to,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Звонящие"
    sheet.sheet_view.showGridLines = False
    sheet.append(["Номер телефона", "Количество звонков", "Тип контакта", "Первый звонок", "Последний звонок", "Первый разговор, сек.", "Общая длительность, сек.", "Квалифицированных", "Внутренние номера"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D3B31")
    for item in result.items:
        sheet.append([
            item.phone_number, item.call_count,
            "Повторное обращение" if item.contact_type == "repeat" else "Первое обращение",
            item.first_call_at.replace(tzinfo=None), item.last_call_at.replace(tzinfo=None),
            item.first_call_duration_seconds, item.total_duration_seconds,
            item.qualified_calls, ", ".join(item.extensions),
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [20, 20, 22, 21, 21, 24, 25, 23, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    output = BytesIO()
    workbook.save(output)
    return Response(
        output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="revora-callers.xlsx"'},
    )

@router.post("/rule-sets", response_model=RuleSetResponse, status_code=status.HTTP_201_CREATED)
async def create_rule_set(payload: RuleSetRequest, user: CurrentUser, session: Session) -> RuleSetResponse:
    return await CallQualityService(session).create_rule_set(user, payload)


@router.get("/calls/{call_id}/analysis", response_model=CallAnalysisResponse)
async def get_analysis(
    call_id: UUID, user: CurrentUser, session: Session
) -> CallAnalysisResponse:
    return await CallQualityService(session).analysis(user, call_id)


@router.post("/calls/{call_id}/reanalyze", response_model=CallAnalysisResponse)
async def reanalyze(
    call_id: UUID, user: CurrentUser, session: Session
) -> CallAnalysisResponse:
    return await CallQualityService(session).reanalyze(user, call_id)


@router.get("/operators", response_model=OperatorPerformanceResponse)
async def operator_performance(
    user: CurrentUser,
    session: Session,
    date_from: date | None = None,
    date_to: date | None = None,
) -> OperatorPerformanceResponse:
    if date_from and date_to and date_from > date_to:
        raise AppError("INVALID_DATE_RANGE", "date_from must not be after date_to", 422)
    return await CallQualityService(session).operator_performance(user, date_from, date_to)


@router.post(
    "/manual-tests",
    response_model=ManualTestResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def manual_test(
    request: Request,
    user: CurrentUser,
    session: Session,
    settings: RuntimeSettings,
) -> ManualTestResponse:
    if user.role != UserRole.OWNER:
        raise AppError("FORBIDDEN", "Only the owner can test call analysis", 403)
    content_type = request.headers.get("content-type", "").split(";", 1)[0].lower()
    if content_type not in ALLOWED_AUDIO_TYPES:
        raise AppError("AUDIO_TYPE_UNSUPPORTED", "Upload MP3, M4A, WAV, OGG or WEBM audio", 415)
    chunks, size = [], 0
    async for chunk in request.stream():
        size += len(chunk)
        if size > settings.call_max_audio_bytes:
            raise AppError("AUDIO_TOO_LARGE", "Audio file exceeds the configured size limit", 413)
        chunks.append(chunk)
    if not size:
        raise AppError("AUDIO_EMPTY", "Audio file is empty", 422)
    filename = Path(unquote(request.headers.get("x-filename", "manual-test.mp3"))).name[:200]
    employee = unquote(request.headers.get("x-operator-name", "")).strip()[:150] or "Ручной тест"
    rules = await ensure_default_rule_set(session, user.tenant_id)
    if rules is None:
        raise AppError("CALL_RULE_SET_MISSING", "A call quality rule set could not be created", 409)
    external_id = f"manual:{uuid4()}"
    call = Call(
        tenant_id=user.tenant_id,
        external_id=external_id,
        phone_hash=sha256(external_id.encode()).hexdigest(),
        direction="manual",
        started_at=datetime.now(UTC),
        duration_seconds=None,
        outcome="manual_test",
        external_user=employee,
        phone_masked=None,
        recording_url=None,
    )
    session.add(call)
    await session.flush()
    analysis = CallQualityAnalysis(
        tenant_id=user.tenant_id,
        call_id=call.id,
        rule_set_id=rules.id,
        status="processing",
        queued_at=datetime.now(UTC),
    )
    session.add(analysis)
    await session.flush()
    await session.commit()
    final_status = await CallQualityPipeline(session, settings).run_inline_audio(
        user.tenant_id,
        analysis.id,
        b"".join(chunks),
        filename=filename,
        content_type=content_type,
    )
    return ManualTestResponse(
        call_id=call.id, analysis_id=analysis.id, status=final_status
    )
