from hashlib import sha256
import hmac
from io import BytesIO

import httpx
import pytest
from openpyxl import Workbook

from app.modules.whatsapp.ai import (
    is_urgent_or_sensitive,
    retrieve_knowledge,
    rules_decision,
)
from app.modules.whatsapp.knowledge_import import import_knowledge_workbook
from app.modules.whatsapp.meta_client import (
    MetaEmbeddedSignupClient,
    MetaEmbeddedSignupError,
)
from app.modules.whatsapp.security import (
    decrypt_contact,
    encrypt_contact,
    valid_meta_signature,
)
from app.modules.whatsapp.schemas import KnowledgeCreateRequest, KnowledgeUpdateRequest


def test_retrieval_returns_only_a_matching_approved_answer() -> None:
    match = retrieve_knowledge(
        "Сколько стоит консультация ортодонта?",
        [
            (
                "Стоимость консультации ортодонта",
                "Консультация ортодонта стоит 10 000 тенге.",
            ),
            ("Как нас найти", "Клиника находится по адресу Абая, 10."),
        ],
    )

    assert match is not None
    assert match.title == "Стоимость консультации ортодонта"
    decision = rules_decision(match)
    assert decision.handoff is False
    assert "10 000" in decision.reply


def test_unknown_question_is_handed_to_a_human() -> None:
    decision = rules_decision(
        retrieve_knowledge(
            "Какой диагноз у меня по фотографии?",
            [("Адрес", "Клиника находится по адресу Абая, 10.")],
        )
    )

    assert decision.handoff is True
    assert decision.confidence == 0


def test_urgent_and_human_requests_are_detected() -> None:
    assert is_urgent_or_sensitive("Кровотечение не останавливается") is not None
    assert is_urgent_or_sensitive("Позовите живого человека") is not None
    assert is_urgent_or_sensitive("Где вы находитесь?") is None


def test_patient_data_encryption_and_meta_signature() -> None:
    key = "clinic-specific-test-key-long-enough"
    ciphertext = encrypt_contact("+77001234567", key)

    assert "+77001234567" not in ciphertext
    assert decrypt_contact(ciphertext, key) == "+77001234567"

    body = b'{"object":"whatsapp_business_account"}'
    secret = "meta-app-secret"
    signature = "sha256=" + hmac.new(secret.encode(), body, sha256).hexdigest()
    assert valid_meta_signature(body, signature, secret) is True
    assert valid_meta_signature(body + b"x", signature, secret) is False


def test_workbook_rows_start_unapproved_and_promotions_stay_human_only() -> None:
    workbook = Workbook()
    faq = workbook.active
    faq.title = "FAQ_КОНСУЛЬТАЦИЯ"
    faq.append(
        [
            "Сколько стоит консультация?",
            "Стоимость первичной консультации составляет 10 000 тенге.",
        ]
    )
    promo = workbook.create_sheet("РАССЫЛКА КП")
    promo.append(
        [
            "Акция июля",
            "Напишите пациентам о специальном предложении до конца месяца.",
        ]
    )
    internal = workbook.create_sheet("Лист14")
    internal.append(["77750001122", "Внутренняя заметка о пациенте, не для бота."])
    data = BytesIO()
    workbook.save(data)

    rows = import_knowledge_workbook(data.getvalue(), "scripts.xlsx")

    assert len(rows) == 2
    assert next(row for row in rows if row.category == "FAQ_КОНСУЛЬТАЦИЯ").risk_level == "review"
    assert next(row for row in rows if row.category == "РАССЫЛКА КП").risk_level == "human_only"


def test_manual_knowledge_requires_an_answer_and_supports_partial_updates() -> None:
    with pytest.raises(ValueError):
        KnowledgeCreateRequest(category="FAQ", title="Цена", content_ru=None, content_kk=None)

    item = KnowledgeCreateRequest(
        category="FAQ",
        title="Стоимость консультации",
        content_ru="Консультация стоит 10 000 тенге.",
    )
    assert item.risk_level == "review"
    update = KnowledgeUpdateRequest(approved=True)
    assert update.approved is True and update.title is None


@pytest.mark.asyncio
async def test_embedded_signup_exchanges_code_and_subscribes_verified_waba() -> None:
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        if request.url.path.endswith("/oauth/access_token"):
            assert request.url.params["client_id"] == "app-123"
            assert request.url.params["client_secret"] == "app-secret"
            assert request.url.params["code"] == "single-use-code"
            return httpx.Response(
                200,
                json={"access_token": "channel-token", "expires_in": 3600},
            )
        if request.url.path.endswith("/waba-123/phone_numbers"):
            assert request.headers["Authorization"] == "Bearer channel-token"
            assert request.url.params["appsecret_proof"]
            return httpx.Response(
                200,
                json={
                    "data": [
                        {
                            "id": "phone-123",
                            "display_phone_number": "+7 777 000 11 22",
                            "verified_name": "Revora Test",
                        }
                    ]
                },
            )
        if request.url.path.endswith("/waba-123/subscribed_apps"):
            return httpx.Response(200, json={"success": True})
        return httpx.Response(404)

    client = MetaEmbeddedSignupClient(
        app_id="app-123",
        app_secret="app-secret",
        graph_version="v25.0",
        transport=httpx.MockTransport(handler),
    )

    token, expires_in = await client.exchange_code("single-use-code")
    phone = await client.verify_and_subscribe(
        token=token,
        waba_id="waba-123",
        phone_number_id="phone-123",
    )

    assert expires_in == 3600
    assert phone["verified_name"] == "Revora Test"
    assert [request.method for request in requests] == ["GET", "GET", "POST"]


@pytest.mark.asyncio
async def test_embedded_signup_rejects_phone_from_another_waba() -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, json={"data": []})

    client = MetaEmbeddedSignupClient(
        app_id="app-123",
        app_secret="app-secret",
        graph_version="v25.0",
        transport=httpx.MockTransport(handler),
    )

    with pytest.raises(MetaEmbeddedSignupError, match="does not belong"):
        await client.verify_and_subscribe(
            token="channel-token",
            waba_id="waba-123",
            phone_number_id="phone-123",
        )
